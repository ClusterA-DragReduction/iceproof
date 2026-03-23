# -*- coding: utf-8 -*-
import streamlit as st
import pandas as pd
import numpy as np
import math
import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter, column_index_from_string
import plotly.graph_objects as go
from plotly.subplots import make_subplots
from wing_drawing import draw_wing_schematic

st.set_page_config(
    page_title="无人机防冰热载荷计算",
    page_icon="❄️",
    layout="wide",
    initial_sidebar_state="expanded",
)

R_air = 287.058
L_fusion = 2499000
cp_air = 1009
Pr = 0.72

import json
import os
import glob
from datetime import datetime

SCHEMES_DIR = "schemes"

def ensure_schemes_dir():
    """确保方案存储目录存在"""
    if not os.path.exists(SCHEMES_DIR):
        os.makedirs(SCHEMES_DIR)

def save_scheme_to_file(scheme_data):
    """保存方案到JSON文件"""
    ensure_schemes_dir()
    # 清理文件名中的非法字符
    name = scheme_data["name"].replace("/", "_").replace("\\", "_")
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{name}_{timestamp}.json"
    filepath = os.path.join(SCHEMES_DIR, filename)
    with open(filepath, "w", encoding="utf-8") as f:
        json.dump(scheme_data, f, ensure_ascii=False, indent=2)
    return filepath

def load_schemes():
    """加载所有已保存的方案，同时返回文件名"""
    ensure_schemes_dir()
    schemes = []
    for filepath in glob.glob(os.path.join(SCHEMES_DIR, "*.json")):
        try:
            with open(filepath, "r", encoding="utf-8") as f:
                data = json.load(f)
                data["filename"] = os.path.basename(filepath)
                schemes.append(data)
        except Exception as e:
            st.warning(f"无法加载方案 {filepath}: {e}")
    schemes.sort(key=lambda x: x.get("timestamp", ""), reverse=True)
    return schemes


def parse_beta_from_string(content):
    data = []
    for line in content.split("\n"):
        line = line.strip()
        if not line or line.startswith("#"):
            continue
        parts = line.split()
        if len(parts) >= 5:
            try:
                x = float(parts[0])
                z = float(parts[2])
                beta = float(parts[4])
                if beta >= 0.001:
                    data.append([x, z, beta])
            except ValueError:
                continue
    return pd.DataFrame(data, columns=["x", "z", "beta"])


def parse_cp_from_string(content):
    data = []
    cp_values = []
    for line in content.split("\n"):
        line = line.strip()
        if not line or line.startswith("#") or line.startswith("X"):
            continue
        parts = line.split(",")
        if len(parts) >= 4:
            try:
                x = float(parts[0].strip())
                z = float(parts[2].strip())
                cp = float(parts[3].strip())
                data.append([x, z, cp])
                cp_values.append(cp)
            except ValueError:
                continue

    df = pd.DataFrame(data, columns=["x", "z", "cp"])

    decimal_places = 6
    for cp in cp_values:
        if not math.isclose(cp, round(cp)):
            s = f"{cp:.10f}".rstrip("0")
            if "." in s:
                dec = len(s.split(".")[1])
                if dec > 0:
                    decimal_places = dec
                    break

    mask = df["cp"] >= 1.0
    if mask.any():
        replacement = round(0.999999, decimal_places)
        df.loc[mask, "cp"] = replacement

    return df


def shift_origin(df):
    dist = np.sqrt(df["x"] ** 2 + df["z"] ** 2)
    idx = dist.idxmin()
    x0, z0 = df.loc[idx, ["x", "z"]]
    df = df.copy()
    df["x"] -= x0
    df["z"] -= z0
    df.loc[idx, "x"] = 0.0
    df.loc[idx, "z"] = 0.0
    return df, idx


def split_surfaces_by_order(df, origin_idx):
    eps = 1e-8
    df.loc[origin_idx, "x"] = 0.0
    df.loc[origin_idx, "z"] = 0.0

    upper = df[df["z"] >= 0].copy()
    lower = df[df["z"] < 0].copy()

    upper["dist"] = np.sqrt(upper["x"] ** 2 + upper["z"] ** 2)
    upper = upper.sort_values("dist").drop(columns="dist").reset_index(drop=True)
    lower["dist"] = np.sqrt(lower["x"] ** 2 + lower["z"] ** 2)
    lower = lower.sort_values("dist").drop(columns="dist").reset_index(drop=True)

    origin_row = df.loc[origin_idx][["x", "z", "beta", "cp"]].copy()
    origin_row["x"] = 0.0
    origin_row["z"] = 0.0

    if not any((abs(lower["x"]) < eps) & (abs(lower["z"]) < eps)):
        lower = pd.concat([pd.DataFrame([origin_row]), lower], ignore_index=True)

    if not (abs(upper.loc[0, "x"]) < eps and abs(upper.loc[0, "z"]) < eps):
        upper = pd.concat([pd.DataFrame([origin_row]), upper], ignore_index=True)

    upper = upper.drop_duplicates(subset=["x", "z"]).reset_index(drop=True)
    lower = lower.drop_duplicates(subset=["x", "z"]).reset_index(drop=True)

    return upper, lower


def compute_surface(df_surface, V_inf, T_inf, T_wall, H_alt, LWC, surface_name):
    n = len(df_surface)
    cols = {
        key: [0.0] * n
        for key in [
            "B",
            "C",
            "D",
            "E",
            "F",
            "G",
            "H",
            "I",
            "J",
            "K",
            "L",
            "M",
            "N",
            "O",
            "P",
            "Q",
            "R",
            "S",
            "T",
            "U",
            "V",
            "W",
            "X",
            "Y",
            "Z",
            "AA",
            "AB",
            "AC",
            "AD",
            "AE",
            "AF",
            "AG",
            "AH",
            "AI",
            "AJ",
            "AK",
            "AL",
            "AM",
            "AN",
        ]
    }

    for i, row in df_surface.iterrows():
        x = row["x"]
        z = row["z"]
        beta = row["beta"]
        cp_val = row["cp"]

        cols["B"][i] = x
        cols["C"][i] = z
        cols["S"][i] = cp_val
        cols["W"][i] = beta
        cols["G"][i] = V_inf
        cols["H"][i] = T_inf
        cols["I"][i] = T_wall
        cols["J"][i] = H_alt
        cols["X"][i] = LWC

        if i == 0:
            cols["D"][i] = 0.0
            cols["E"][i] = 0.0
        else:
            dx = x - cols["B"][i - 1]
            dz = z - cols["C"][i - 1]
            ds = math.sqrt(dx * dx + dz * dz)
            cols["D"][i] = ds
            cols["E"][i] = cols["E"][i - 1] + ds

    for i in range(n):
        G = cols["G"][i]
        H = cols["H"][i]
        I_val = cols["I"][i]
        J = cols["J"][i]
        X = cols["X"][i]
        cp_val = cols["S"][i]
        beta = cols["W"][i]
        E = cols["E"][i]

        K = 101325 * (1 - 0.0000225577 * J) ** 5.255
        cols["K"][i] = K
        L = K / (R_air * H)
        cols["L"][i] = L
        M = 0.0241 * (H / 273.15) ** 0.84
        cols["M"][i] = M
        mu = 1.716e-5 * (H / 273.15) ** 1.5 * (273.15 + 110.4) / (H + 110.4)
        N_val = mu / L
        cols["N"][i] = N_val
        F = G * math.sqrt(max(0, 1 - cp_val))
        cols["F"][i] = F
        O_val = F**1.87
        cols["O"][i] = O_val

        if i == 0:
            cols["P"][i] = 0.0
        else:
            O_prev = cols["O"][i - 1]
            O_curr = O_val
            E_prev = cols["E"][i - 1]
            E_curr = E
            cols["P"][i] = cols["P"][i - 1] + (O_curr + O_prev) / 2 * (E_curr - E_prev)

        if N_val > 0:
            Q = 0.296 * M / math.sqrt(N_val)
        else:
            Q = 0.0
        cols["Q"][i] = Q

        if F > 0 and cols["P"][i] > 0:
            R = Q * (F**-2.87 * cols["P"][i]) ** (-0.5)
        else:
            R = 0.0
        cols["R"][i] = R

        T = H + (G * G - F * F) / (2 * cp_air)
        cols["T"][i] = T
        U = R * (I_val - T)
        cols["U"][i] = U
        V = 0.843 * R * F * F / (2 * cp_air)
        cols["V"][i] = V
        Y = 0.1 * G * beta * X * 4687 * (I_val - H)
        cols["Y"][i] = Y
        cols["Z"][i] = Y * 10

        Twc = I_val - 273.15
        AA = 609.603 + 49.495 * Twc + 1.739 * Twc**2 + 0.031 * Twc**3 + 0.0002 * Twc**4
        cols["AA"][i] = AA
        Tic = H - 273.15
        AB = 609.603 + 49.495 * Tic + 1.739 * Tic**2 + 0.031 * Tic**3 + 0.0002 * Tic**4
        cols["AB"][i] = AB
        AC = 0.5 * cp_val * 0.589 * G * G + K
        cols["AC"][i] = AC

        if AC > 0 and K > 0:
            AD = 0.685 * (R * L_fusion / cp_air) * ((AA / AC) - (AB / K)) * 0.1
        else:
            AD = 0.0
        cols["AD"][i] = AD
        cols["AE"][i] = AD * 10

        AF = 0.5 * beta * X * G**3
        cols["AF"][i] = AF

        cols["AG"][i] = 1000 * E
        cols["AH"][i] = beta
        cols["AI"][i] = U + Y + AD - V - AF
        cols["AJ"][i] = U
        cols["AK"][i] = AD
        cols["AL"][i] = Y
        cols["AM"][i] = AF
        cols["AN"][i] = V

    result = pd.DataFrame(cols)
    result["surface"] = surface_name
    return result


def extract_summary_data(upper_df, lower_df):
    def extract_rows(df, reverse=False, negate_s=False):
        rows = []
        eps = 1e-6
        for _, row in df.iterrows():
            s = row["AG"]
            if abs(s) < eps:
                continue
            if negate_s:
                s = -s
            hydro = {
                "s": s,
                "beta": row["AH"],
                "Qn": row["AI"],
                "Qa": row["U"],
                "Qe": row["AK"],
                "Qw": row["AL"],
                "Qwv": row["AM"],
                "Qv": row["AN"],
            }
            non_hydro = {
                "s": s,
                "beta": row["AH"],
                "Qn": row["U"] + row["Z"] + row["AE"] - row["V"] - row["AF"],
                "Qa": row["U"],
                "Qe": row["AE"],
                "Qw": row["Z"],
                "Qwv": row["AF"],
                "Qv": row["V"],
            }
            rows.append((hydro, non_hydro))
        rows.sort(key=lambda x: x[0]["s"], reverse=reverse)
        return rows

    upper_rows = extract_rows(upper_df, reverse=True)
    lower_rows = extract_rows(lower_df, reverse=False, negate_s=True)

    all_rows = []
    for hydro, non_hydro in upper_rows + lower_rows:
        all_rows.append(("hydro", hydro))
        all_rows.append(("non_hydro", non_hydro))

    all_rows.sort(key=lambda x: x[1]["s"], reverse=True)

    all_rows_filtered = []
    for typ, vals in all_rows:
        if any(
            v is None or (isinstance(v, float) and math.isnan(v)) for v in vals.values()
        ):
            continue
        all_rows_filtered.append((typ, vals))

    hydro_data = []
    non_hydro_data = []
    for typ, vals in all_rows_filtered:
        if typ == "hydro":
            hydro_data.append(vals)
        else:
            non_hydro_data.append(vals)

    return pd.DataFrame(hydro_data), pd.DataFrame(non_hydro_data)


def calculate_zones(hydro_df, non_hydro_df):
    # 计算全局最大beta（用于分区阈值）
    max_beta_hydro = hydro_df["beta"].max() if len(hydro_df) > 0 else 0
    max_beta_non = non_hydro_df["beta"].max() if len(non_hydro_df) > 0 else 0
    max_beta_global = max(max_beta_hydro, max_beta_non)

    A_threshold = max_beta_global * 0.7
    B_threshold = max_beta_global * 0.4

    def get_zone_info(df):
        zones = {"A": [], "B": [], "C": [], "D": [], "E": []}
        for _, row in df.iterrows():
            s = row["s"]
            beta = row["beta"]
            qn = row["Qn"]
            if beta >= A_threshold:
                zones["A"].append((s, qn))
            elif beta >= B_threshold:
                if s >= 0:
                    zones["B"].append((s, qn))
                else:
                    zones["C"].append((s, qn))
            else:
                if s >= 0:
                    zones["D"].append((s, qn))
                else:
                    zones["E"].append((s, qn))

        result = {}
        for zone, points in zones.items():
            if not points:
                result[zone] = {
                    "s_min": 0,
                    "s_max": 0,
                    "width": 0,
                    "typical_qn": 0,
                    "max_qn": 0,
                }
                continue
            s_vals = [p[0] for p in points]
            qn_vals = [p[1] for p in points]
            sorted_qn = sorted(qn_vals, reverse=True)
            result[zone] = {
                "s_min": min(s_vals),
                "s_max": max(s_vals),
                "width": max(s_vals) - min(s_vals),
                "typical_qn": sorted_qn[2] if len(sorted_qn) >= 3 else (sorted_qn[-1] if sorted_qn else 0),
                "max_qn": max(qn_vals) if qn_vals else 0,
            }
        return result

    hydro_zones = get_zone_info(hydro_df)
    non_hydro_zones = get_zone_info(non_hydro_df)
    return hydro_zones, non_hydro_zones


def create_excel_output(upper_df, lower_df, V_inf, T_inf, T_wall, H_alt, LWC):
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "01"

    header_font = Font(bold=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    fill_red = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    fill_green = PatternFill(
        start_color="90EE90", end_color="90EE90", fill_type="solid"
    )
    fill_yellow = PatternFill(
        start_color="FFFF00", end_color="FFFF00", fill_type="solid"
    )
    fill_none = PatternFill(fill_type=None)

    ws1["A2"] = "上表面"
    ws1["A2"].font = header_font
    ws1["AM2"] = "疏水"
    ws1["AM2"].font = header_font
    ws1["AQ2"] = "不疏水"
    ws1["AQ2"].font = header_font

    headers = [
        "点",
        "X (m)",
        "Z (m)",
        "ΔS (m)",
        "累计弧长 S (m)",
        "vl",
        "来流速度",
        "来流温度",
        "防冰温度",
        "飞行高度",
        "大气压",
        "空气密度",
        "热导率",
        "运动粘度",
        "系数",
        "梯形数值积分",
        "系数",
        "换热系数Htc",
        "压力系数",
        "TL",
        "01_Qa对流换热",
        "02_Qv摩擦生热",
        "beta",
        "LWC",
        "03_Qw",
        "03_Qw（不加涂层）",
        "es",
        "ea",
        "pl",
        "04_Qe蒸发比热流",
        "04_Qe(不加涂层）",
        "05_Qwv水滴动能比热流",
        "s",
        "β",
        "Qn",
        "Qa",
        "Qe",
        "Qw",
        "Qwv",
        "Qv",
        "",
        "",
        "s",
        "β",
        "Qn",
        "Qa",
        "Qe",
        "Qw",
        "Qwv",
        "Qv",
    ]
    for col, header in enumerate(headers, start=1):
        cell = ws1.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.border = thin_border
        if header in ["来流速度", "来流温度", "防冰温度", "飞行高度", "LWC"]:
            cell.fill = fill_red
        if header in ["X (m)", "Z (m)", "压力系数", "beta"]:
            cell.fill = fill_yellow

    for col in range(1, 51):
        ws1.column_dimensions[get_column_letter(col)].width = 12

    start_row = 4
    current_row = start_row

    def get_row_values(row):
        return {
            k: row[k]
            for k in [
                "B",
                "C",
                "D",
                "E",
                "F",
                "G",
                "H",
                "I",
                "J",
                "K",
                "L",
                "M",
                "N",
                "O",
                "P",
                "Q",
                "R",
                "S",
                "T",
                "U",
                "V",
                "W",
                "X",
                "Y",
                "Z",
                "AA",
                "AB",
                "AC",
                "AD",
                "AE",
                "AF",
                "AG",
                "AH",
                "AI",
                "AJ",
                "AK",
                "AL",
                "AM",
                "AN",
            ]
        }

    for _, row in upper_df.iterrows():
        ws1.cell(row=current_row, column=1, value=current_row - start_row)
        vals = get_row_values(row)
        for col_letter in [
            "B",
            "C",
            "D",
            "E",
            "F",
            "G",
            "H",
            "I",
            "J",
            "K",
            "L",
            "M",
            "N",
            "O",
            "P",
            "Q",
            "R",
            "S",
            "T",
            "U",
            "V",
            "W",
            "X",
            "Y",
            "Z",
            "AA",
            "AB",
            "AC",
            "AD",
            "AE",
            "AF",
            "AG",
            "AH",
            "AI",
            "AJ",
            "AK",
            "AL",
            "AM",
            "AN",
        ]:
            col_idx = column_index_from_string(col_letter)
            val = vals[col_letter]
            cell = ws1.cell(row=current_row, column=col_idx, value=val)
            if col_letter in ["G", "H", "I", "J", "X"]:
                cell.fill = fill_red
            elif col_letter in ["B", "C", "S", "W"]:
                cell.fill = fill_yellow
            elif col_letter in ["AG", "AH", "AI", "AJ", "AK", "AL", "AM", "AN"]:
                cell.fill = fill_green
            else:
                cell.fill = fill_none
            if col_letter not in ["B", "C", "D", "E", "S", "W", "X"]:
                cell.number_format = "0.00"
        for col, val in [
            (43, vals["AG"]),
            (44, vals["AH"]),
            (
                45,
                vals["AJ"]
                + vals["AK"] * 10
                + vals["AL"] * 10
                - vals["AM"]
                - vals["AN"],
            ),
            (46, vals["AJ"]),
            (47, vals["AK"] * 10),
            (48, vals["AL"] * 10),
            (49, vals["AM"]),
            (50, vals["AN"]),
        ]:
            cell = ws1.cell(row=current_row, column=col, value=val)
            cell.fill = fill_green
            cell.number_format = "0.00"
        current_row += 1

    ws1.cell(row=current_row, column=1, value="下表面")
    ws1.cell(row=current_row, column=1).font = Font(bold=True)
    current_row += 1

    for _, row in lower_df.iterrows():
        ws1.cell(row=current_row, column=1, value=current_row - start_row)
        vals = get_row_values(row)
        for col_letter in [
            "B",
            "C",
            "D",
            "E",
            "F",
            "G",
            "H",
            "I",
            "J",
            "K",
            "L",
            "M",
            "N",
            "O",
            "P",
            "Q",
            "R",
            "S",
            "T",
            "U",
            "V",
            "W",
            "X",
            "Y",
            "Z",
            "AA",
            "AB",
            "AC",
            "AD",
            "AE",
            "AF",
            "AG",
            "AH",
            "AI",
            "AJ",
            "AK",
            "AL",
            "AM",
            "AN",
        ]:
            col_idx = column_index_from_string(col_letter)
            val = vals[col_letter]
            cell = ws1.cell(row=current_row, column=col_idx, value=val)
            if col_letter in ["G", "H", "I", "J", "X"]:
                cell.fill = fill_red
            elif col_letter in ["B", "C", "S", "W"]:
                cell.fill = fill_yellow
            elif col_letter in ["AG", "AH", "AI", "AJ", "AK", "AL", "AM", "AN"]:
                cell.fill = fill_green
            else:
                cell.fill = fill_none
            if col_letter not in ["B", "C", "D", "E", "S", "W", "X"]:
                cell.number_format = "0.00"
        for col, val in [
            (43, vals["AG"]),
            (44, vals["AH"]),
            (
                45,
                vals["AJ"]
                + vals["AK"] * 10
                + vals["AL"] * 10
                - vals["AM"]
                - vals["AN"],
            ),
            (46, vals["AJ"]),
            (47, vals["AK"] * 10),
            (48, vals["AL"] * 10),
            (49, vals["AM"]),
            (50, vals["AN"]),
        ]:
            cell = ws1.cell(row=current_row, column=col, value=val)
            cell.fill = fill_green
            cell.number_format = "0.00"
        current_row += 1

    for col in range(1, 51):
        ws1.column_dimensions[get_column_letter(col)].width = 12

    ws2 = wb.create_sheet("01汇总")

    ws2["B1"] = "疏水"
    ws2["N1"] = "不疏水"
    ws2["B1"].font = header_font
    ws2["N1"].font = header_font

    zone_colors = {
        "A": PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),
        "B": PatternFill(start_color="FF6600", end_color="FF6600", fill_type="solid"),
        "C": PatternFill(start_color="0066FF", end_color="0066FF", fill_type="solid"),
        "D": PatternFill(start_color="00CC00", end_color="00CC00", fill_type="solid"),
        "E": PatternFill(start_color="00CC00", end_color="00CC00", fill_type="solid"),
    }

    summary_headers = ["分区", "s[mm]", "β", "Qn", "Qa", "Qe", "Qw", "Qwv", "Qv"]
    for idx, header in enumerate(summary_headers, start=1):
        cell = ws2.cell(row=2, column=idx, value=header)
        cell.font = header_font

    summary_headers_right = ["分区", "s[mm]", "β", "Qn", "Qa", "Qe", "Qw", "Qwv", "Qv"]
    for idx, header in enumerate(summary_headers_right, start=14):
        cell = ws2.cell(row=2, column=idx, value=header)
        cell.font = header_font

    hydro_df_export, non_hydro_df_export = extract_summary_data(upper_df, lower_df)

    def get_zone(s, beta, max_beta):
        if s >= 0:
            if beta >= max_beta * 0.7:
                return "A"
            elif beta >= max_beta * 0.4:
                return "B"
            else:
                return "D"
        else:
            if beta >= max_beta * 0.7:
                return "A"
            elif beta >= max_beta * 0.4:
                return "C"
            else:
                return "E"

    max_beta_hydro = hydro_df_export["beta"].max() if len(hydro_df_export) > 0 else 0
    max_beta_non = (
        non_hydro_df_export["beta"].max() if len(non_hydro_df_export) > 0 else 0
    )

    start_row = 3
    for i in range(len(hydro_df_export)):
        excel_row = start_row + i
        row_data = hydro_df_export.iloc[i]

        zone = get_zone(row_data["s"], row_data["beta"], max_beta_hydro)

        ws2.cell(row=excel_row, column=1, value=zone)
        ws2.cell(row=excel_row, column=2, value=row_data["s"]).number_format = "0.00"
        ws2.cell(row=excel_row, column=3, value=row_data["beta"]).number_format = "0.00"
        ws2.cell(row=excel_row, column=4, value=row_data["Qn"]).number_format = "0.00"
        ws2.cell(row=excel_row, column=5, value=row_data["Qa"]).number_format = "0.00"
        ws2.cell(row=excel_row, column=6, value=row_data["Qe"]).number_format = "0.00"
        ws2.cell(row=excel_row, column=7, value=row_data["Qw"]).number_format = "0.00"
        ws2.cell(row=excel_row, column=8, value=row_data["Qwv"]).number_format = "0.00"
        ws2.cell(row=excel_row, column=9, value=row_data["Qv"]).number_format = "0.00"

        for col in range(1, 10):
            ws2.cell(row=excel_row, column=col).fill = zone_colors.get(zone, fill_none)

    for i in range(len(non_hydro_df_export)):
        excel_row = start_row + i
        row_data = non_hydro_df_export.iloc[i]

        zone = get_zone(row_data["s"], row_data["beta"], max_beta_non)

        ws2.cell(row=excel_row, column=14, value=zone)
        ws2.cell(row=excel_row, column=15, value=row_data["s"]).number_format = "0.00"
        ws2.cell(
            row=excel_row, column=16, value=row_data["beta"]
        ).number_format = "0.00"
        ws2.cell(row=excel_row, column=17, value=row_data["Qn"]).number_format = "0.00"
        ws2.cell(row=excel_row, column=18, value=row_data["Qa"]).number_format = "0.00"
        ws2.cell(row=excel_row, column=19, value=row_data["Qe"]).number_format = "0.00"
        ws2.cell(row=excel_row, column=20, value=row_data["Qw"]).number_format = "0.00"
        ws2.cell(row=excel_row, column=21, value=row_data["Qwv"]).number_format = "0.00"
        ws2.cell(row=excel_row, column=22, value=row_data["Qv"]).number_format = "0.00"

        for col in range(14, 23):
            ws2.cell(row=excel_row, column=col).fill = zone_colors.get(zone, fill_none)

    for col in range(1, 23):
        ws2.column_dimensions[get_column_letter(col)].width = 12

    return wb


def main():
    st.markdown(
        """
    <style>
    .main-header {
    font-size: 2.1rem !important;   /* 比原来的2.2rem略大，也可改为3rem等 */
    font-weight: bold;
    color: #1E88E5;
    text-align: center;
    margin-bottom: 1.5rem;
    padding-bottom: 0.8rem;
    border-bottom: 3px solid #1E88E5;
}
    .sub-header {
        font-size: 1.1rem;
        font-weight: bold;
        color: #424242;
        margin-top: 1rem;
        margin-bottom: 0.8rem;
    }
    .success-box {
        background-color: #E8F5E9;
        padding: 0.8rem;
        border-radius: 0.5rem;
        border-left: 4px solid #4CAF50;
        font-size: 0.9rem;
    }
    .info-box {
        background-color: #E3F2FD;
        padding: 0.8rem;
        border-radius: 0.5rem;
        border-left: 4px solid #2196F3;
        font-size: 0.9rem;
    }
    /* 方案管理卡片 */
.schemes-card {
    background-color: #F0F7FF;
    padding: 1rem 1rem 0.5rem 1rem;
    border-radius: 0.5rem;
    border-left: 4px solid #2196F3;
    margin-bottom: 1rem;
}
.schemes-card h4 {
    margin-top: 0;
    margin-bottom: 0.5rem;
    color: #0B5E7E;
}
/* 方案列表项之间的分隔线样式 */
div[data-testid="stVerticalBlock"] > div:has(hr) {
    margin: 0.5rem 0;
}
    div[data-testid="stMetricValue"] {
        font-size: 1.2rem !important;
    }
    div[data-testid="stMetricLabel"] {
        font-size: 0.9rem !important;
    }
    div[data-testid="stMetricDelta"] {
        font-size: 0.8rem !important;
    }
    div[data-testid="column"] {
        padding: 0 0.5rem;
    }
    
    </style>
    """,
        unsafe_allow_html=True,
    )

    st.markdown(
        '<p class="main-header">❄️ 无人机电热防冰热载荷计算系统</p>',
        unsafe_allow_html=True,
    )

    if "calculation_done" not in st.session_state:
        st.session_state.calculation_done = False
    if "results" not in st.session_state:
        st.session_state.results = None

    # 侧边栏模式选择
    with st.sidebar:
        st.markdown("## 🎯 工作模式")
        mode = st.radio(
            "请选择工作模式",
            options=["热载荷计算", "分区设计工具"],
            index=0,
            help="热载荷计算：上传文件、设置参数，计算防冰功率并保存方案；分区设计工具：加载已有方案，生成设计图纸。"
        )
        st.markdown("---")

        if mode == "热载荷计算":
            st.markdown("## 📊 输入参数")
            st.markdown("---")

            st.markdown("### 📁 数据文件")
            beta_file = st.file_uploader("上传 Beta 文件", type=None)
            cp_file = st.file_uploader("上传 Cp 文件", type=None)

            st.markdown("---")
            st.markdown("### 🔧 飞行参数")

            col1, col2 = st.columns(2)
            with col1:
                V_inf = st.number_input("来流速度 (m/s)", value=41.7, step=0.1, format="%.1f")
            with col2:
                T_inf = st.number_input("来流温度 (K)", value=263.15, step=0.1, format="%.2f")

            col3, col4 = st.columns(2)
            with col3:
                T_wall = st.number_input("防冰温度 (K)", value=276.15, step=0.1, format="%.2f")
            with col4:
                H_alt = st.number_input("飞行高度 (m)", value=4000.0, step=100.0, format="%.0f")

            LWC = st.number_input(
                "液态水含量 LWC (kg/m³)",
                value=0.00034,
                step=0.00001,
                format="%.6f",
            )

            st.markdown("---")
            st.markdown("### 🛩️ 机体参数")
            wing_span = st.number_input("翼展单边长度 (m)", value=4.9, step=0.1, format="%.1f")

            st.markdown("---")
            calculate_btn = st.button("🚀 开始计算", type="primary", use_container_width=True)

            st.markdown("---")
            st.markdown("""
            ### ℹ️ 使用说明
            1. 上传 Beta 和 Cp 数据文件
            - 支持任意格式或无扩展名

            2. 设置飞行参数和机体参数

            3. 点击计算按钮
            - 查看数据表格和图表
            - 下载Excel结果文件
            """)

        elif mode == "分区设计工具":
            # 设计工具模式下侧边栏可放一些说明或全局参数（可选）
            st.markdown("### 🛠️ 设计工具说明")
            st.markdown("""
            - 选择已保存的分区方案
            - 设置电偶位置、凸台宽度等参数
            - 生成工程图纸
            """)

    # 根据模式显示主要内容
    if mode == "热载荷计算":
        # 如果文件未上传，显示方案管理界面
        if not beta_file or not cp_file:
            # 统一的欢迎和引导卡片
            st.markdown(
                """
                <div class="info-box">
                    <h5>👋 欢迎使用</h5>
                    <p>请在左侧上传 <b>Beta</b> 和 <b>Cp</b> 数据文件，然后设置参数，点击计算按钮开始分析。</p>
                </div>
                """,
                unsafe_allow_html=True,
            )

            schemes = load_schemes()
            if not schemes:
                st.info("暂无保存的方案。请上传文件并完成计算后保存方案。")
            else:
                # 方案管理卡片
                st.markdown(
                    """
                    <div class="schemes-card">
                        <h6>📁 已保存的分区方案</h6>
                        <p style="margin-bottom: 1rem;">您可以直接删除不需要的方案，或继续上传文件进行计算。</p>
                    </div>
                    """,
                    unsafe_allow_html=True,
                )

                # 构建方案列表
                for i, s in enumerate(schemes):
                    with st.container():
                        col1, col2, col3, col4, col5, col6 = st.columns([2.5, 1.5, 1.2, 1.8, 2.5, 0.6])
                        with col1:
                            st.markdown(f"**{s['name']}**")
                        with col2:
                            st.write(f"{s['wing_span']} m")
                        with col3:
                            st.write(s['scheme_type'])
                        with col4:
                            st.write(s['timestamp'][:19])
                        with col5:
                            widths_str = ", ".join([f"{k}:{v:.0f}" for k, v in s['widths'].items() if v > 0])
                            st.write(widths_str)
                        with col6:
                            if st.button("🗑️", key=f"del_{i}", help="删除此方案"):
                                filepath = os.path.join(SCHEMES_DIR, s["filename"])
                                os.remove(filepath)
                                st.success(f"方案「{s['name']}」已删除")
                                st.rerun()
                        st.markdown("---")

            # 无需重复的“上传文件”提示，欢迎消息已包含
            return

        # 文件已上传，显示参数确认和计算按钮（按钮已在侧边栏，但需在下方处理）
        st.markdown("##### 📄 已上传文件")
        # 可展示文件基本信息
        try:
            beta_content = beta_file.getvalue().decode("utf-8")
            cp_content = cp_file.getvalue().decode("utf-8")

            beta_df = parse_beta_from_string(beta_content)
            cp_df = parse_cp_from_string(cp_content)

            # ========== 新增格式校验 ==========
            if beta_df.empty:
                st.error("❌ Beta 文件格式错误：未能解析出有效的 Beta 数据点。\n"
                         "请确保文件每行包含至少5个空格分隔的字段，且第1、3、5列分别为 x, z, beta（beta ≥ 0.001）。")
                return

            if cp_df.empty:
                st.error("❌ Cp 文件格式错误：未能解析出有效的 Cp 数据点。\n"
                         "请确保文件每行包含至少4个逗号分隔的字段，且第1、3、4列分别为 x, z, cp（cp ≤ 1）。")
                return

            # 可选：数据点数量警告
            if len(beta_df) < 3:
                st.warning("⚠️ Beta 数据点少于 3 个，计算精度可能受影响，但可继续。")
            if len(cp_df) < 3:
                st.warning("⚠️ Cp 数据点少于 3 个，计算精度可能受影响，但可继续。")
            # ========== 格式校验结束 ==========

            # 初始化折叠状态和滚动标志
            if "data_preview_collapsed" not in st.session_state:
                st.session_state.data_preview_collapsed = False
            if "scroll_to_results" not in st.session_state:
                st.session_state.scroll_to_results = False

            with st.expander(
                    f"📈 数据加载成功 - Beta: {len(beta_df)} 点, Cp: {len(cp_df)} 点",
                    expanded=not st.session_state.data_preview_collapsed,
            ):
                col1, col2 = st.columns(2)
                with col1:
                    st.write("**Beta 数据预览**")
                    st.dataframe(beta_df.head(10), use_container_width=True)
                with col2:
                    st.write("**Cp 数据预览**")
                    st.dataframe(cp_df.head(10), use_container_width=True)
        except Exception as e:
            st.error(f"❌ 文件解析错误: {str(e)}")
            return

        if calculate_btn:
            # 折叠数据预览（将在本次运行中生效）
            st.session_state.data_preview_collapsed = True

            with st.spinner("🧮 正在进行热载荷计算..."):
                try:
                    beta_df, origin_idx = shift_origin(beta_df)
                    cp_df, _ = shift_origin(cp_df)

                    from scipy.spatial import KDTree

                    tree = KDTree(cp_df[["x", "z"]])
                    _, idx = tree.query(beta_df[["x", "z"]])
                    matched_df = beta_df.copy()
                    matched_df["cp"] = cp_df.iloc[idx]["cp"].values

                    upper, lower = split_surfaces_by_order(matched_df, origin_idx)

                    upper_calc = compute_surface(
                        upper, V_inf, T_inf, T_wall, H_alt, LWC, "upper"
                    )
                    lower_calc = compute_surface(
                        lower, V_inf, T_inf, T_wall, H_alt, LWC, "lower"
                    )

                    hydro_df, non_hydro_df = extract_summary_data(upper_calc, lower_calc)

                    # zones = calculate_zones(hydro_df, non_hydro_df)
                    hydro_zones, non_hydro_zones = calculate_zones(hydro_df, non_hydro_df)

                    st.session_state.results = {
                        "upper_calc": upper_calc,
                        "lower_calc": lower_calc,
                        "hydro_df": hydro_df,
                        "non_hydro_df": non_hydro_df,
                        "hydro_zones": hydro_zones,
                        "non_hydro_zones": non_hydro_zones,
                        # "zones": zones,
                        "params": {
                            "V_inf": V_inf,
                            "T_inf": T_inf,
                            "T_wall": T_wall,
                            "H_alt": H_alt,
                            "LWC": LWC,
                            "wing_span": wing_span,
                        },
                    }
                    st.session_state.calculation_done = True

                except Exception as e:
                    st.error(f"❌ 计算过程中发生错误: {str(e)}")
                    import traceback

                    st.code(traceback.format_exc())
                    return

        if st.session_state.calculation_done and st.session_state.results:
            # 添加锚点
            st.markdown('<div id="result-anchor"></div>', unsafe_allow_html=True)
            # 滚动到锚点
            st.components.v1.html("""
                <script>
                    document.getElementById('result-anchor').scrollIntoView({behavior: 'smooth'});
                </script>
            """, height=0)
            results = st.session_state.results
            hydro_df = results["hydro_df"]
            non_hydro_df = results["non_hydro_df"]
            # zones = results["zones"]
            params = results["params"]

            st.markdown("---")
            st.markdown('<p class="sub-header">✅ 计算完成</p>', unsafe_allow_html=True)

            st.markdown(
                f"""
                <div class="success-box">
                    <b>计算参数:</b> 速度={params["V_inf"]} m/s, 来流温度={params["T_inf"]} K, 
                    防冰温度={params["T_wall"]} K, 高度={params["H_alt"]} m, LWC={params["LWC"]} kg/m³,
                    翼展单边={params["wing_span"]} m
                </div>
                """,
                unsafe_allow_html=True,
            )

            # st.subheader("📊 关键热载荷指标")
            st.markdown("##### 📉 关键热载荷指标")
            col1, col2, col3, col4 = st.columns(4)

            with col1:
                max_Qn = hydro_df["Qn"].max() if len(hydro_df) > 0 else 0
                st.metric("最大总热流 Qn (疏水)", f"{max_Qn:.2f} W/m²")
            with col2:
                max_Qa = hydro_df["Qa"].max() if len(hydro_df) > 0 else 0
                st.metric("最大对流换热 Qa", f"{max_Qa:.2f} W/m²")
            with col3:
                max_Qe = hydro_df["Qe"].max() if len(hydro_df) > 0 else 0
                st.metric("最大蒸发换热 Qe", f"{max_Qe:.2f} W/m²")
            with col4:
                max_Qw = hydro_df["Qw"].max() if len(hydro_df) > 0 else 0
                st.metric("最大水滴换热 Qw", f"{max_Qw:.2f} W/m²")

            # st.subheader("📉 热载荷分布曲线")
            st.markdown("##### 📉 热载荷分布曲线")
            # tab1, tab2, tab3, tab4 = st.tabs(
            #     ["热流变化曲线", "表面形状", "防冰功率密度分布", "数据表格"]
            # )
            # 在结果展示部分，构建选项卡名称列表
            tab_names = ["热流变化曲线", "表面形状", "防冰功率密度分布", "数据表格"]
            if load_schemes():
                tab_names.append("📐 分区设计工具")
            tabs = st.tabs(tab_names)


            with tabs[0]:
                if len(hydro_df) > 0:
                    # 准备 Htc 数据（从上下表面合并）
                    upper = results["upper_calc"]
                    lower = results["lower_calc"]
                    htc_data = pd.concat([
                        upper[["AG", "R"]].rename(columns={"AG": "s", "R": "Htc"}),
                        lower[["AG", "R"]].rename(columns={"AG": "s", "R": "Htc"})
                    ])
                    htc_data = htc_data.sort_values("s").drop_duplicates(subset=["s"]).reset_index(drop=True)

                    # 创建两行两列的子图
                    fig = make_subplots(
                        rows=2, cols=2,
                        subplot_titles=("总热流 Qn 分布", "各分项热流分布", "β 角度分布", "局部换热系数 Htc 分布"),
                        column_widths=[0.5, 0.5],
                        row_heights=[0.5, 0.5],  # 两行等高
                        vertical_spacing=0.25,
                        horizontal_spacing=0.15,
                        specs=[
                            [{"type": "scatter"}, {"type": "scatter"}],
                            [{"type": "scatter"}, {"type": "scatter"}]
                        ],
                    )

                    # ---- 子图1：总热流 Qn ----
                    fig.add_trace(
                        go.Scatter(
                            x=hydro_df["s"],
                            y=hydro_df["Qn"],
                            mode="lines+markers",
                            name="疏水 Qn",
                            line=dict(color="#2196F3", width=2),
                            marker=dict(size=4, opacity=0.6),
                        ),
                        row=1, col=1,
                    )
                    if len(non_hydro_df) > 0:
                        fig.add_trace(
                            go.Scatter(
                                x=non_hydro_df["s"],
                                y=non_hydro_df["Qn"],
                                mode="lines+markers",
                                name="不疏水 Qn",
                                line=dict(color="#FF5722", width=2),
                                marker=dict(size=4, opacity=0.6),
                            ),
                            row=1, col=1,
                        )

                    # ---- 子图2：各分项热流 ----
                    fig.add_trace(
                        go.Scatter(
                            x=hydro_df["s"],
                            y=hydro_df["Qa"],
                            mode="lines",
                            name="Qa 对流",
                            line=dict(color="#4CAF50", width=2),
                        ),
                        row=1, col=2,
                    )
                    fig.add_trace(
                        go.Scatter(
                            x=hydro_df["s"],
                            y=hydro_df["Qe"],
                            mode="lines",
                            name="Qe 蒸发",
                            line=dict(color="#FF9800", width=2),
                        ),
                        row=1, col=2,
                    )
                    fig.add_trace(
                        go.Scatter(
                            x=hydro_df["s"],
                            y=hydro_df["Qw"],
                            mode="lines",
                            name="Qw 水滴",
                            line=dict(color="#9C27B0", width=2),
                        ),
                        row=1, col=2,
                    )
                    fig.add_trace(
                        go.Scatter(
                            x=hydro_df["s"],
                            y=hydro_df["Qwv"],
                            mode="lines",
                            name="Qwv 动能",
                            line=dict(color="#E91E63", width=2),
                        ),
                        row=1, col=2,
                    )
                    fig.add_trace(
                        go.Scatter(
                            x=hydro_df["s"],
                            y=hydro_df["Qv"],
                            mode="lines",
                            name="Qv 摩擦",
                            line=dict(color="#607D8B", width=2),
                        ),
                        row=1, col=2,
                    )

                    # ---- 子图3：β 角度分布 ----
                    max_beta_idx = hydro_df["beta"].idxmax()
                    max_beta_val = hydro_df.loc[max_beta_idx, "beta"]
                    max_beta_s = hydro_df.loc[max_beta_idx, "s"]

                    fig.add_trace(
                        go.Scatter(
                            x=hydro_df["s"],
                            y=hydro_df["beta"],
                            mode="lines+markers",
                            name="β 角度",
                            line=dict(color="#3F51B5", width=2.5),
                            marker=dict(size=4, color="#3F51B5", opacity=0.5),
                        ),
                        row=2, col=1,
                    )
                    # 最大值点
                    fig.add_trace(
                        go.Scatter(
                            x=[max_beta_s],
                            y=[max_beta_val],
                            mode="markers",
                            name=f"β max = {max_beta_val:.3f}",
                            marker=dict(color="red", size=12, symbol="circle", line=dict(width=1, color="darkred")),
                        ),
                        row=2, col=1,
                    )
                    # 数值标注
                    fig.add_annotation(
                        x=max_beta_s,
                        y=max_beta_val,
                        text=f"{max_beta_val:.3f}",
                        showarrow=True,
                        arrowhead=2,
                        ax=20,
                        ay=-30,
                        font=dict(size=11, color="red", weight="bold"),
                        row=2, col=1,
                    )

                    # ---- 子图4：局部换热系数 Htc ----
                    # 标记最大值点
                    max_htc_idx = htc_data["Htc"].idxmax()
                    max_htc_val = htc_data.loc[max_htc_idx, "Htc"]
                    max_htc_s = htc_data.loc[max_htc_idx, "s"]

                    fig.add_trace(
                        go.Scatter(
                            x=htc_data["s"],
                            y=htc_data["Htc"],
                            mode="lines+markers",
                            name="Htc 换热系数",
                            line=dict(color="#00ACC1", width=2.5),
                            marker=dict(size=4, color="#00ACC1", opacity=0.5),
                            hovertemplate='弧长: %{x:.1f} mm<br>Htc: %{y:.1f} W/m²K<br>换热系数越大，散热越快<extra></extra>'

                        ),
                        row=2, col=2,
                    )
                    # 最大值点
                    fig.add_trace(
                        go.Scatter(
                            x=[max_htc_s],
                            y=[max_htc_val],
                            mode="markers",
                            name=f"Htc max = {max_htc_val:.1f}",
                            marker=dict(color="darkorange", size=12, symbol="circle",
                                        line=dict(width=1, color="brown")),
                        ),
                        row=2, col=2,
                    )
                    # 数值标注
                    fig.add_annotation(
                        x=max_htc_s,
                        y=max_htc_val,
                        text=f"{max_htc_val:.1f}",
                        showarrow=True,
                        arrowhead=2,
                        ax=20,
                        ay=-30,
                        font=dict(size=11, color="darkorange", weight="bold"),
                        row=2, col=2,
                    )

                    # 更新布局：图例放在顶部居中，提高位置避免遮挡
                    fig.update_layout(
                        height=700,  # 适当增加高度
                        showlegend=True,
                        legend=dict(
                            orientation="h",
                            yanchor="bottom",
                            y=1.05,
                            xanchor="center",
                            x=0.5,
                            bgcolor="rgba(255,255,255,0.8)",
                            bordercolor="LightGray",
                            borderwidth=1,
                            font=dict(size=9.5),
                            itemclick="toggle",
                            itemdoubleclick="toggleothers",
                        ),
                        margin=dict(l=40, r=40, t=110, b=40),
                    )

                    # 设置坐标轴标签
                    fig.update_xaxes(title_text="弧长 s [mm]", row=1, col=1)
                    fig.update_xaxes(title_text="弧长 s [mm]", row=1, col=2)
                    fig.update_xaxes(title_text="弧长 s [mm]", row=2, col=1)
                    fig.update_xaxes(title_text="弧长 s [mm]", row=2, col=2)
                    fig.update_yaxes(title_text="热流 [W/m²]", row=1, col=1)
                    fig.update_yaxes(title_text="热流 [W/m²]", row=1, col=2)
                    fig.update_yaxes(title_text="β [rad]", row=2, col=1)
                    fig.update_yaxes(title_text="Htc [W/m²K]", row=2, col=2)

                    # 调整子图标题字体
                    for i in range(1, 5):
                        fig.layout.annotations[i - 1].update(font=dict(size=11))

                    st.plotly_chart(fig, use_container_width=True)
                    # st.caption("局部换热系数 Htc 表示对流换热能力，值越大，所需防冰功率越高。")
            # st.caption(
            #     """
            #     **局部换热系数 Htc 分布说明**
            #     - Htc 表示对流换热系数（单位：W/m²K），反映空气流过表面时的冷却能力。
            #     - 值越大，表明散热越快，所需防冰加热功率越高。
            #     - 通常在驻点（β 最大）附近 Htc 达到峰值，沿流向逐渐降低。
            #     - 结合 β 和热流分量，可用于优化分区加热设计。
            #     """
            # )
            with tabs[1]:
                upper = results["upper_calc"]
                lower = results["lower_calc"]

                fig2 = go.Figure()

                fig2.add_trace(
                    go.Scatter(
                        x=upper["B"],
                        y=upper["C"],
                        mode="lines+markers",
                        name="上表面",
                        line=dict(color="#2196F3", width=2),
                        marker=dict(size=6),
                    )
                )

                fig2.add_trace(
                    go.Scatter(
                        x=lower["B"],
                        y=lower["C"],
                        mode="lines+markers",
                        name="下表面",
                        line=dict(color="#FF5722", width=2),
                        marker=dict(size=6),
                    )
                )

                fig2.update_layout(
                    title="机翼表面形状",
                    xaxis_title="X (m)",
                    yaxis_title="Z (m)",
                    height=500,
                    legend=dict(
                        orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1
                    ),
                )

                st.plotly_chart(fig2, use_container_width=True)

            with tabs[2]:
                st.markdown("##### 防冰功率密度大小分布 (s vs β)")

                if len(hydro_df) > 0:
                    fig3 = go.Figure()

                    max_beta = hydro_df["beta"].max()
                    A_threshold = max_beta * 0.7
                    B_threshold = max_beta * 0.4

                    hydro_sorted = hydro_df.sort_values("s")

                    for zone_name, zone_color, zone_range in [
                        ("A区", "#FF0000", (A_threshold, max_beta)),
                        ("B区", "#FF6600", (B_threshold, A_threshold)),
                        ("D区", "#00CC00", (0, B_threshold)),
                    ]:
                        zone_data = hydro_sorted[
                            (hydro_sorted["beta"] >= zone_range[0])
                            & (hydro_sorted["beta"] < zone_range[1])
                            ]
                        fig3.add_trace(
                            go.Scatter(
                                x=zone_data["s"],
                                y=zone_data["beta"],
                                mode="lines+markers",
                                name=zone_name,
                                line=dict(color=zone_color, width=2),
                                marker=dict(size=6),
                            )
                        )

                    if len(non_hydro_df) > 0:
                        non_hydro_sorted = non_hydro_df.sort_values("s")
                        for zone_name, zone_color, zone_range in [
                            ("C区", "#0066FF", (B_threshold, A_threshold)),
                            ("E区", "#00CC00", (0, B_threshold)),
                        ]:
                            zone_data = non_hydro_sorted[
                                (non_hydro_sorted["beta"] >= zone_range[0])
                                & (non_hydro_sorted["beta"] < zone_range[1])
                                ]
                            fig3.add_trace(
                                go.Scatter(
                                    x=zone_data["s"],
                                    y=zone_data["beta"],
                                    mode="lines+markers",
                                    name=zone_name,
                                    line=dict(color=zone_color, width=2),
                                    marker=dict(size=6),
                                )
                            )

                    fig3.update_layout(
                        title="防冰功率密度大小分布 (s vs β)",
                        xaxis_title="弧长 s [mm]",
                        yaxis_title="β [rad]",
                        height=500,
                        legend=dict(
                            orientation="h",
                            yanchor="bottom",
                            y=1.02,
                            xanchor="right",
                            x=1,
                        ),
                    )

                    st.plotly_chart(fig3, use_container_width=True)

            with tabs[3]:
                col1, col2 = st.columns(2)

                with col1:
                    st.markdown("##### 疏水表面热流数据")
                    if len(hydro_df) > 0:
                        display_df = hydro_df.copy().round(4)
                        st.dataframe(display_df, use_container_width=True, height=400)
                    else:
                        st.warning("无数据")

                with col2:
                    st.markdown("##### 不疏水表面热流数据")
                    if len(non_hydro_df) > 0:
                        display_df2 = non_hydro_df.copy().round(4)
                        st.dataframe(display_df2, use_container_width=True, height=400)
                    else:
                        st.warning("无数据")

            # 如果存在第五个选项卡
            if len(tabs) > 4:
                with tabs[4]:
                    st.markdown("## ✏️ 主机翼分区设计工具")
                    st.info("选择之前保存的分区方案，并配置电偶等参数，生成设计图纸。")
                    schemes = load_schemes()
                    if not schemes:
                        st.warning("暂无保存的分区方案。请先保存一个方案。")
                    else:
                        # 方案选择和删除按钮（两列）
                        col_sel, col_del = st.columns([4, 1])
                        with col_sel:
                            selected_index = st.selectbox(
                                "选择分区方案",
                                options=range(len(schemes)),
                                format_func=lambda i: f"{schemes[i]['name']} ({schemes[i]['timestamp'][:19]})",
                                key="design_tab_scheme"
                            )
                        with col_del:
                            if st.button("🗑️ 删除当前方案", key="delete_scheme_tab"):
                                filepath = os.path.join(SCHEMES_DIR, schemes[selected_index]["filename"])
                                os.remove(filepath)
                                st.success("方案已删除")
                                st.rerun()  # 刷新页面更新方案列表
                        selected_scheme = schemes[selected_index]

                        # 显示方案信息
                        col_info1, col_info2 = st.columns(2)
                        with col_info1:
                            st.metric("翼展单边长度", f"{selected_scheme['wing_span']} m")
                        with col_info2:
                            st.metric("方案类型", selected_scheme['scheme_type'])

                        widths_df = pd.DataFrame([selected_scheme["widths"]], index=["宽度 (mm)"]).T
                        st.dataframe(widths_df, use_container_width=True)

                        # 设计参数表单（避免参数调整时页面刷新）
                        with st.form("design_params_form_tab"):
                            st.markdown("### ⚙️ 设计参数设置")
                            col1, col2 = st.columns(2)
                            with col1:
                                total_length = st.number_input("主体总长 (mm)", value=500.0, step=10.0,
                                                               key="total_length_tab")
                                bottom_height = st.number_input("底部段高度 (mm)", value=72.5, step=5.0,
                                                                key="bottom_height_tab")
                                top_height = st.number_input("顶部段高度 (mm)", value=72.5, step=5.0,
                                                             key="top_height_tab")
                                boss_width = st.number_input("右侧凸台宽度 (mm)", value=10.0, step=2.0,
                                                             key="boss_width_tab")
                                left_boss_width = st.number_input("左侧黄色矩形宽度 (mm)", value=10.0, step=2.0,
                                                                  key="left_boss_width_tab")
                                hole_diameter = st.number_input("圆孔直径 (mm)", value=6.0, step=1.0,
                                                                key="hole_diameter_tab")
                            with col2:
                                thermocouple_width = st.number_input("电偶宽度 (mm)", value=7.0, step=1.0,
                                                                     key="thermocouple_width_tab")
                                thermocouple_extension = st.number_input("电偶超出顶部距离 (mm)", value=10.0, step=5.0,
                                                                         key="thermocouple_extension_tab")
                                left_thermocouple_dist = st.number_input("左电偶距左边框距离 (mm)", value=20.0,
                                                                         step=5.0, key="left_thermocouple_dist_tab")
                                right_thermocouple_dist = st.number_input("右电偶距凸台左侧距离 (mm)", value=10.0,
                                                                          step=5.0, key="right_thermocouple_dist_tab")

                            submitted = st.form_submit_button("生成设计图纸", type="primary")
                            if submitted:
                                with st.spinner("生成图纸中..."):
                                    wing_span_mm = selected_scheme["wing_span"] * 1000
                                    widths = selected_scheme["widths"]

                                    fig_no = draw_wing_schematic(
                                        widths=widths,
                                        wing_span=wing_span_mm,
                                        total_length=total_length,
                                        bottom_height=bottom_height,
                                        top_height=top_height,
                                        boss_width=boss_width,
                                        left_boss_width=left_boss_width,
                                        hole_diameter=hole_diameter,
                                        show_thermocouple=False,
                                    )
                                    fig_left = draw_wing_schematic(
                                        widths=widths,
                                        wing_span=wing_span_mm,
                                        total_length=total_length,
                                        bottom_height=bottom_height,
                                        top_height=top_height,
                                        boss_width=boss_width,
                                        left_boss_width=left_boss_width,
                                        hole_diameter=hole_diameter,
                                        show_thermocouple=True,
                                        thermocouple_side='left',
                                        thermocouple_width=thermocouple_width,
                                        thermocouple_extension=thermocouple_extension,
                                        left_thermocouple_dist=left_thermocouple_dist,
                                        right_thermocouple_dist=right_thermocouple_dist,
                                    )
                                    fig_right = draw_wing_schematic(
                                        widths=widths,
                                        wing_span=wing_span_mm,
                                        total_length=total_length,
                                        bottom_height=bottom_height,
                                        top_height=top_height,
                                        boss_width=boss_width,
                                        left_boss_width=left_boss_width,
                                        hole_diameter=hole_diameter,
                                        show_thermocouple=True,
                                        thermocouple_side='right',
                                        thermocouple_width=thermocouple_width,
                                        thermocouple_extension=thermocouple_extension,
                                        left_thermocouple_dist=left_thermocouple_dist,
                                        right_thermocouple_dist=right_thermocouple_dist,
                                    )

                                    # 保存图形和字节数据到 session_state（使用特定键名）
                                    st.session_state.design_figs_tab = {
                                        'no': fig_no,
                                        'left': fig_left,
                                        'right': fig_right
                                    }
                                    buf_no = io.BytesIO()
                                    fig_no.savefig(buf_no, format="png", dpi=300, bbox_inches='tight')
                                    st.session_state.design_buf_no_tab = buf_no.getvalue()

                                    buf_left = io.BytesIO()
                                    fig_left.savefig(buf_left, format="png", dpi=300, bbox_inches='tight')
                                    st.session_state.design_buf_left_tab = buf_left.getvalue()

                                    buf_right = io.BytesIO()
                                    fig_right.savefig(buf_right, format="png", dpi=300, bbox_inches='tight')
                                    st.session_state.design_buf_right_tab = buf_right.getvalue()

                                    st.success("图纸生成完成！可在下方下载。")

                        # 显示图片和下载按钮（在表单外部，不会因下载而刷新）
                        if 'design_buf_no_tab' in st.session_state:
                            st.markdown("### 🖼️ 设计图纸预览")
                            col_img1, col_img2, col_img3 = st.columns(3)
                            with col_img1:
                                st.pyplot(st.session_state.design_figs_tab['no'])
                            with col_img2:
                                st.pyplot(st.session_state.design_figs_tab['left'])
                            with col_img3:
                                st.pyplot(st.session_state.design_figs_tab['right'])

                            st.markdown("### 📥 下载设计图纸")
                            col_dl1, col_dl2, col_dl3 = st.columns(3)
                            with col_dl1:
                                st.download_button(
                                    label="📸 下载无电偶图",
                                    data=st.session_state.design_buf_no_tab,
                                    file_name="no_thermocouple.png",
                                    mime="image/png"
                                )
                            with col_dl2:
                                st.download_button(
                                    label="📸 下载左电偶图",
                                    data=st.session_state.design_buf_left_tab,
                                    file_name="left_thermocouple.png",
                                    mime="image/png"
                                )
                            with col_dl3:
                                st.download_button(
                                    label="📸 下载右电偶图",
                                    data=st.session_state.design_buf_right_tab,
                                    file_name="right_thermocouple.png",
                                    mime="image/png"
                                )

            # ========== 分区防冰功率计算（可编辑宽度和系数） ==========
            # ========== 分区防冰功率计算（可编辑宽度、Qn和系数） ==========
            st.markdown("##### 🔥 分区防冰功率计算")

            import math

            # 从 results 中获取分区数据
            hydro_zones = results["hydro_zones"]
            non_hydro_zones = results["non_hydro_zones"]

            # 选择查看的方案
            scheme = st.radio("选择防冰方案", ["疏水方案", "不疏水方案"], horizontal=True, key="scheme_radio")
            current_zones = hydro_zones if scheme == "疏水方案" else non_hydro_zones

            # 默认系数（硬编码，与之前保持一致）
            DEFAULT_COEFFS = {
                "A": 1.46,
                "B": 1.34,
                "C": 1.34,
                "D": 1.28,
                "E": 1.28,
            }

            # 分区顺序：与示意图一致（从上到下）
            all_zones = ["C", "B", "A", "D", "E"]

            # 初始化 session_state 中存储编辑宽度、Qn和系数的字典
            if "edited_widths" not in st.session_state:
                st.session_state.edited_widths = {"hydro": {}, "non_hydro": {}}
            if "edited_qns" not in st.session_state:
                st.session_state.edited_qns = {"hydro": {}, "non_hydro": {}}
            if "edited_coeffs" not in st.session_state:
                st.session_state.edited_coeffs = {"hydro": {}, "non_hydro": {}}

            # 获取当前方案的默认值
            default_widths = {}
            default_qns = {}
            for zone in all_zones:
                info = current_zones.get(zone, {"width": 0, "typical_qn": 0})
                raw_width = info["width"]
                default_widths[zone] = math.ceil(raw_width) if raw_width > 0 else 0
                default_qns[zone] = info["typical_qn"]

            default_coeffs = {zone: DEFAULT_COEFFS[zone] for zone in all_zones}

            key = "hydro" if scheme == "疏水方案" else "non_hydro"

            # 初始化或补齐宽度
            if not st.session_state.edited_widths[key]:
                st.session_state.edited_widths[key] = default_widths.copy()
            else:
                for zone in all_zones:
                    if zone not in st.session_state.edited_widths[key]:
                        st.session_state.edited_widths[key][zone] = default_widths.get(zone, 0)

            # 初始化或补齐Qn
            if not st.session_state.edited_qns[key]:
                st.session_state.edited_qns[key] = default_qns.copy()
            else:
                for zone in all_zones:
                    if zone not in st.session_state.edited_qns[key]:
                        st.session_state.edited_qns[key][zone] = default_qns.get(zone, 0)

            # 初始化或补齐系数
            if not st.session_state.edited_coeffs[key]:
                st.session_state.edited_coeffs[key] = default_coeffs.copy()
            else:
                for zone in all_zones:
                    if zone not in st.session_state.edited_coeffs[key]:
                        st.session_state.edited_coeffs[key][zone] = default_coeffs.get(zone, DEFAULT_COEFFS[zone])

            # 提示信息
            st.info("💡 **可编辑列**：您可以直接修改“宽度 (mm)”、“典型Qn (W/m²)”和“修正系数 C”，其他列将自动计算更新。")

            # 构建可编辑数据表（始终显示5个分区）
            edit_data = []
            for zone in all_zones:
                width = st.session_state.edited_widths[key][zone]
                qn = st.session_state.edited_qns[key][zone]
                coeff = st.session_state.edited_coeffs[key][zone]
                qn_modified = qn * coeff
                power = params["wing_span"] * 0.001 * qn_modified * width if width > 0 else 0
                edit_data.append({
                    "分区": zone,
                    "宽度 (mm)": width,
                    "典型Qn (W/m²)": round(qn, 1),
                    "修正系数 C": coeff,
                    "修正后Qn (W/m²)": round(qn_modified, 1),
                    "功率 (W)": round(power, 1),
                })

            edit_df = pd.DataFrame(edit_data)

            # 使用 data_editor 允许编辑宽度、Qn和系数
            edited_df = st.data_editor(
                edit_df,
                column_config={
                    "分区": st.column_config.TextColumn(
                        "分区",
                        disabled=True,
                        help="分区标识（不可编辑）"
                    ),
                    "宽度 (mm)": st.column_config.NumberColumn(
                        "宽度 (mm)",
                        min_value=0,
                        step=1,
                        format="%d",
                        help="可编辑：分区宽度（毫米），修改后功率自动更新"
                    ),
                    "典型Qn (W/m²)": st.column_config.NumberColumn(
                        "典型Qn (W/m²)",
                        min_value=0.0,
                        step=0.1,
                        format="%.1f",
                        help="可编辑：分区典型总热流，修改后修正Qn和功率自动更新"
                    ),
                    "修正系数 C": st.column_config.NumberColumn(
                        "修正系数 C",
                        min_value=0.0,
                        step=0.01,
                        format="%.2f",
                        help="可编辑：模型修正系数，修改后修正Qn和功率自动更新"
                    ),
                    "修正后Qn (W/m²)": st.column_config.NumberColumn(
                        "修正后Qn (W/m²)",
                        disabled=True,
                        format="%.1f",
                        help="修正后Qn = 典型Qn × 修正系数"
                    ),
                    "功率 (W)": st.column_config.NumberColumn(
                        "功率 (W)",
                        disabled=True,
                        format="%.1f",
                        help="功率 = 翼展单边 × 修正后Qn × 宽度 / 1000"
                    ),
                },
                hide_index=True,
                use_container_width=True,
            )

            # 更新 session_state 中的编辑值
            for i, row in edited_df.iterrows():
                zone = row["分区"]
                st.session_state.edited_widths[key][zone] = int(row["宽度 (mm)"])
                st.session_state.edited_qns[key][zone] = row["典型Qn (W/m²)"]
                st.session_state.edited_coeffs[key][zone] = row["修正系数 C"]

            # 重新计算当前方案总功率（使用编辑后值）
            total_power = 0
            for i, row in edited_df.iterrows():
                total_power += row["功率 (W)"]

            st.markdown(
                f"""
                <div style="background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); 
                            padding: 1rem; border-radius: 8px; color: white; text-align: center; margin-top: 1rem;">
                    <h5>🎯 当前方案（{scheme}）防冰总功率: {total_power:.1f} W ({total_power / 1000:.2f} kW)</h5>
                </div>
                """,
                unsafe_allow_html=True,
            )

            # ========== 分区设计示意图（根据编辑后的宽度绘制） ==========
            st.markdown("##### 🛩️ 主机翼防冰分区设计示意图")
            zone_colors_dict = {
                "A": "#FF1493", "B": "#90EE90", "C": "#48D1CC", "D": "#9370DB", "E": "#800080"
            }

            fig_zones = go.Figure()
            # 固定从上到下的顺序：E, D, A, B, C （绘制顺序从下往上，但视觉上C在最上）
            fixed_order = ["E", "D", "A", "B", "C"]
            zone_heights = []
            for zone in fixed_order:
                width = st.session_state.edited_widths[key].get(zone, 0)
                if width > 0:
                    zone_heights.append((zone, width))

            if not zone_heights:
                zone_heights = [("C", 50), ("B", 40), ("A", 30), ("D", 20), ("E", 10)]

            current_y = 0
            for i, (zone, width_mm) in enumerate(zone_heights):
                height = width_mm * 1.5
                y_start = current_y
                y_end = current_y + height

                if zone == "A":
                    label = f"{zone}区：前缘防冰区域"
                elif zone in ["B", "C"]:
                    label = f"{zone}区-上翼面防冰区域"
                else:
                    label = f"{zone}区-下翼面防冰区域"

                fig_zones.add_trace(go.Scatter(
                    x=[0.3, 0.8, 0.8, 0.3, 0.3],
                    y=[y_start, y_start, y_end, y_end, y_start],
                    fill="toself",
                    fillcolor=zone_colors_dict.get(zone, "#888888"),
                    opacity=0.85,
                    line=dict(color="DarkGray", width=1),
                    mode="lines",
                    showlegend=False,
                ))
                fig_zones.add_annotation(
                    x=0.55,
                    y=(y_start + y_end) / 2,
                    text=label,
                    showarrow=False,
                    font=dict(size=11, color="black", weight="bold"),
                )

                # 右侧宽度标注
                arrow_x = 0.82
                mid_y = (y_start + y_end) / 2
                fig_zones.add_shape(type="line", x0=arrow_x - 0.018, y0=y_start, x1=arrow_x + 0.01, y1=y_start,
                                    line=dict(color="#333", width=1))
                fig_zones.add_shape(type="line", x0=arrow_x - 0.018, y0=y_end, x1=arrow_x + 0.01, y1=y_end,
                                    line=dict(color="#333", width=1))
                fig_zones.add_shape(type="line", x0=arrow_x, y0=y_start, x1=arrow_x, y1=y_end,
                                    line=dict(color="#333", width=1))
                fig_zones.add_annotation(
                    x=arrow_x, y=mid_y,
                    text=f"{round(width_mm)}mm",
                    showarrow=False,
                    font=dict(size=10, color="#333"),
                    bgcolor="white",
                )
                current_y = y_end

            total_height = current_y
            zone_total_mm = sum(w for _, w in zone_heights)
            wing_span_mm = int(params["wing_span"] * 1000)

            # 左侧总高度标注
            left_x = 0.28
            fig_zones.add_shape(type="line", x0=left_x - 0.015, y0=0, x1=left_x + 0.015, y1=0,
                                line=dict(color="#333", width=1))
            fig_zones.add_shape(type="line", x0=left_x - 0.015, y0=total_height, x1=left_x + 0.015, y1=total_height,
                                line=dict(color="#333", width=1))
            fig_zones.add_shape(type="line", x0=left_x, y0=0, x1=left_x, y1=total_height,
                                line=dict(color="#333", width=1))
            fig_zones.add_annotation(x=left_x, y=total_height / 2, text=f"{round(zone_total_mm)}mm",
                                     showarrow=False, font=dict(size=10, color="#333"), bgcolor="white")

            # 底部翼展标注
            fig_zones.add_shape(type="line", x0=0.3, y0=-5.5, x1=0.8, y1=-5.5, line=dict(color="#333", width=1))
            fig_zones.add_annotation(x=0.55, y=-5.1, text=f"{wing_span_mm}mm",
                                     showarrow=False, font=dict(size=12, color="#333"), bgcolor="white")

            fig_zones.update_layout(
                title=f"主机翼防冰分区设计示意图（{scheme}）",
                height=max(550, total_height + 180),
                showlegend=False,
                xaxis=dict(showticklabels=False, showgrid=False, zeroline=False, showline=False, range=[0.15, 0.95]),
                yaxis=dict(showticklabels=False, showgrid=False, zeroline=False, showline=False,
                           range=[-10, total_height + 25]),
                plot_bgcolor="white",
                margin=dict(l=20, r=20, t=80, b=100),
            )
            st.plotly_chart(fig_zones, use_container_width=True)

            st.markdown("##### 📥 下载结果")

            wb = create_excel_output(
                results["upper_calc"],
                results["lower_calc"],
                params["V_inf"],
                params["T_inf"],
                params["T_wall"],
                params["H_alt"],
                params["LWC"],
            )

            excel_buffer = io.BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)

            col_d1, col_d2 = st.columns(2)

            with col_d1:
                st.download_button(
                    label="📊 下载完整 Excel 报告",
                    data=excel_buffer,
                    file_name=f"防冰热载荷计算_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    type="primary",
                    use_container_width=True,
                )

            with col_d2:
                csv_buffer = io.StringIO()
                hydro_df.to_csv(csv_buffer, index=False)
                csv_data = csv_buffer.getvalue()

                st.download_button(
                    label="📋 下载疏水数据 (CSV)",
                    data=csv_data,
                    file_name=f"疏水热流数据_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
                    mime="text/csv",
                    use_container_width=True,
                )

            with st.expander("📊 详细统计信息"):
                col1, col2 = st.columns(2)

                with col1:
                    st.write("**疏水表面统计**")
                    if len(hydro_df) > 0:
                        st.dataframe(hydro_df.describe().round(4), use_container_width=True)

                with col2:
                    st.write("**不疏水表面统计**")
                    if len(non_hydro_df) > 0:
                        st.dataframe(
                            non_hydro_df.describe().round(4), use_container_width=True
                        )

            # 保存分区方案（供设计工具使用）
            st.markdown("---")
            st.markdown("##### 💾 保存分区方案（供设计工具使用）")
            st.info("您可以编辑上方的分区宽度后，将当前方案保存，供后续设计工具使用。")

            col_save1, col_save2 = st.columns([3, 1])
            with col_save1:
                scheme_name = st.text_input(
                    "方案名称",
                    value=f"方案_{len(load_schemes()) + 1}",
                    key="scheme_name_input"
                )
            with col_save2:
                if st.button("保存当前方案", type="primary"):
                    key = "hydro" if scheme == "疏水方案" else "non_hydro"
                    widths = st.session_state.edited_widths[key].copy()
                    for zone in ["C", "B", "A", "D", "E"]:
                        if zone not in widths:
                            widths[zone] = 0
                    scheme_data = {
                        "name": scheme_name,
                        "wing_span": params["wing_span"],
                        "widths": widths,
                        "scheme_type": scheme,
                        "timestamp": datetime.now().isoformat()
                    }
                    filepath = save_scheme_to_file(scheme_data)
                    st.toast(f"✅ 方案「{scheme_name}」已保存", icon="💾")

    elif mode == "分区设计工具":
        # 强制页面滚动到顶部
        st.components.v1.html(
            """
            <script>
                window.scrollTo(0, 0);
            </script>
            """,
            height=0,
        )
        st.markdown("##### ✏️ 主机翼分区设计工具")
        st.info("选择之前保存的分区方案，并配置电偶等参数，生成设计图纸。")

        schemes = load_schemes()
        if not schemes:
            st.warning("暂无保存的分区方案。请先切换到「热载荷计算」模式完成计算并保存一个方案。")
        else:
            # 方案选择（无删除按钮）
            selected_index = st.selectbox(
                "选择分区方案",
                options=range(len(schemes)),
                format_func=lambda i: f"{schemes[i]['name']} ({schemes[i]['timestamp'][:19]})",
                key="design_mode_scheme"
            )
            selected_scheme = schemes[selected_index]

            # 显示方案信息
            col_info1, col_info2 = st.columns(2)
            with col_info1:
                st.metric("翼展单边长度", f"{selected_scheme['wing_span']} m")
            with col_info2:
                st.metric("方案类型", selected_scheme['scheme_type'])

            # 使用紧凑的指标卡片显示各分区宽度
            col1, col2, col3, col4, col5 = st.columns(5)
            with col1:
                st.metric("C 区", f"{selected_scheme['widths'].get('C', 0):.0f} mm")
            with col2:
                st.metric("B 区", f"{selected_scheme['widths'].get('B', 0):.0f} mm")
            with col3:
                st.metric("A 区", f"{selected_scheme['widths'].get('A', 0):.0f} mm")
            with col4:
                st.metric("D 区", f"{selected_scheme['widths'].get('D', 0):.0f} mm")
            with col5:
                st.metric("E 区", f"{selected_scheme['widths'].get('E', 0):.0f} mm")

            # 设计参数表单（避免参数调整时页面刷新）
            with st.form("design_params_form_mode"):
                st.markdown("##### ⚙️ 设计参数设置")
                col1, col2 = st.columns(2)
                with col1:
                    total_length = st.number_input("主体总长 (mm)", value=500.0, step=10.0, key="total_length_mode")
                    bottom_height = st.number_input("底部段高度 (mm)", value=72.5, step=5.0, key="bottom_height_mode")
                    top_height = st.number_input("顶部段高度 (mm)", value=72.5, step=5.0, key="top_height_mode")
                    boss_width = st.number_input("右侧凸台宽度 (mm)", value=10.0, step=2.0, key="boss_width_mode")
                    left_boss_width = st.number_input("左侧黄色矩形宽度 (mm)", value=10.0, step=2.0, key="left_boss_width_mode")
                    hole_diameter = st.number_input("圆孔直径 (mm)", value=6.0, step=1.0, key="hole_diameter_mode")
                with col2:
                    thermocouple_width = st.number_input("电偶宽度 (mm)", value=7.0, step=1.0, key="thermocouple_width_mode")
                    thermocouple_extension = st.number_input("电偶超出顶部距离 (mm)", value=10.0, step=5.0, key="thermocouple_extension_mode")
                    left_thermocouple_dist = st.number_input("左电偶距左边框距离 (mm)", value=20.0, step=5.0, key="left_thermocouple_dist_mode")
                    right_thermocouple_dist = st.number_input("右电偶距凸台左侧距离 (mm)", value=10.0, step=5.0, key="right_thermocouple_dist_mode")

                submitted = st.form_submit_button("生成设计图纸", type="primary")
                if submitted:
                    with st.spinner("生成图纸中..."):
                        wing_span_mm = selected_scheme["wing_span"] * 1000
                        widths = selected_scheme["widths"]

                        fig_no = draw_wing_schematic(
                            widths=widths,
                            wing_span=wing_span_mm,
                            total_length=total_length,
                            bottom_height=bottom_height,
                            top_height=top_height,
                            boss_width=boss_width,
                            left_boss_width=left_boss_width,
                            hole_diameter=hole_diameter,
                            show_thermocouple=False,
                        )
                        fig_left = draw_wing_schematic(
                            widths=widths,
                            wing_span=wing_span_mm,
                            total_length=total_length,
                            bottom_height=bottom_height,
                            top_height=top_height,
                            boss_width=boss_width,
                            left_boss_width=left_boss_width,
                            hole_diameter=hole_diameter,
                            show_thermocouple=True,
                            thermocouple_side='left',
                            thermocouple_width=thermocouple_width,
                            thermocouple_extension=thermocouple_extension,
                            left_thermocouple_dist=left_thermocouple_dist,
                            right_thermocouple_dist=right_thermocouple_dist,
                        )
                        fig_right = draw_wing_schematic(
                            widths=widths,
                            wing_span=wing_span_mm,
                            total_length=total_length,
                            bottom_height=bottom_height,
                            top_height=top_height,
                            boss_width=boss_width,
                            left_boss_width=left_boss_width,
                            hole_diameter=hole_diameter,
                            show_thermocouple=True,
                            thermocouple_side='right',
                            thermocouple_width=thermocouple_width,
                            thermocouple_extension=thermocouple_extension,
                            left_thermocouple_dist=left_thermocouple_dist,
                            right_thermocouple_dist=right_thermocouple_dist,
                        )

                        # 保存到 session_state（使用模式专用键名）
                        st.session_state.design_figs_mode = {
                            'no': fig_no,
                            'left': fig_left,
                            'right': fig_right
                        }
                        buf_no = io.BytesIO()
                        fig_no.savefig(buf_no, format="png", dpi=300, bbox_inches='tight')
                        st.session_state.design_buf_no_mode = buf_no.getvalue()

                        buf_left = io.BytesIO()
                        fig_left.savefig(buf_left, format="png", dpi=300, bbox_inches='tight')
                        st.session_state.design_buf_left_mode = buf_left.getvalue()

                        buf_right = io.BytesIO()
                        fig_right.savefig(buf_right, format="png", dpi=300, bbox_inches='tight')
                        st.session_state.design_buf_right_mode = buf_right.getvalue()

                        st.success("图纸生成完成！可在下方下载。")

            # 显示图片和下载按钮（在表单外部，不会因下载而刷新）
            if 'design_buf_no_mode' in st.session_state:
                st.markdown("### 🖼️ 设计图纸预览")
                col_img1, col_img2, col_img3 = st.columns(3)
                with col_img1:
                    st.pyplot(st.session_state.design_figs_mode['no'])
                with col_img2:
                    st.pyplot(st.session_state.design_figs_mode['left'])
                with col_img3:
                    st.pyplot(st.session_state.design_figs_mode['right'])

                st.markdown("### 📥 下载设计图纸")
                col_dl1, col_dl2, col_dl3 = st.columns(3)
                with col_dl1:
                    st.download_button(
                        label="📸 下载无电偶图",
                        data=st.session_state.design_buf_no_mode,
                        file_name="no_thermocouple.png",
                        mime="image/png"
                    )
                with col_dl2:
                    st.download_button(
                        label="📸 下载左电偶图",
                        data=st.session_state.design_buf_left_mode,
                        file_name="left_thermocouple.png",
                        mime="image/png"
                    )
                with col_dl3:
                    st.download_button(
                        label="📸 下载右电偶图",
                        data=st.session_state.design_buf_right_mode,
                        file_name="right_thermocouple.png",
                        mime="image/png"
                    )

if __name__ == "__main__":
    main()
